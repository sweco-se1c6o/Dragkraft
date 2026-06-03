from __future__ import annotations

import argparse
import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Prepare a temp MATLAB legacy run directory with an xlsread shim."
    )
    parser.add_argument("--sheet", default="NyProfil")
    parser.add_argument(
        "--temp-dir",
        type=Path,
        default=Path(tempfile.gettempdir()) / "dragkraft_matlab_baseline",
    )
    args = parser.parse_args()

    repo_root = Path(__file__).resolve().parents[1]
    legacy_dir = repo_root / "old"
    args.temp_dir.mkdir(parents=True, exist_ok=True)
    for path in legacy_dir.glob("*.m"):
        shutil.copy2(path, args.temp_dir / path.name)
    shutil.copy2(legacy_dir / "luleaHamn3.xlsx", args.temp_dir / "luleaHamn3.xlsx")
    _write_xlsread_shim(
        workbook_path=legacy_dir / "luleaHamn3.xlsx",
        sheet_name=args.sheet,
        target_path=args.temp_dir / "xlsread.m",
    )

    output_dir = repo_root / "tests" / "fixtures" / "matlab_nyprofil_default"
    print("Prepared MATLAB baseline temp directory:")
    print(args.temp_dir)
    print()
    print("Run MATLAB export with:")
    print(
        "matlab -wait -nosplash -nodesktop "
        f"-r \"addpath('{repo_root / 'tools'}'); "
        f"export_matlab_baseline('{output_dir}', '{args.temp_dir}'); exit\""
    )
    return 0


def _write_xlsread_shim(
    *,
    workbook_path: Path,
    sheet_name: str,
    target_path: Path,
) -> None:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook[sheet_name]
    max_col = 35
    lines = [
        "function [num, txt, raw] = xlsread(varargin)",
        "% Auto-generated shim for Dragkraft baseline export; bypasses Excel COM.",
        "num = [];",
        "txt = {};",
        "raw = { ...",
    ]
    max_row = _legacy_contract_row_count(worksheet)
    for row_number in range(1, max_row + 1):
        values = [
            _matlab_value(worksheet.cell(row_number, column).value)
            for column in range(1, max_col + 1)
        ]
        suffix = "; ..." if row_number < worksheet.max_row else " ..."
        lines.append("    " + ", ".join(values) + suffix)
    lines.extend(["};", "end", ""])
    target_path.write_text("\n".join(lines), encoding="utf-8")
    workbook.close()


def _legacy_contract_row_count(worksheet) -> int:
    primary_columns = (2, 7, 12, 17, 20, 25, 30)
    last_needed = 5
    for column in primary_columns:
        row_number = 5
        while worksheet.cell(row_number, column).value is not None:
            row_number += 1
        last_needed = max(last_needed, row_number)
    return last_needed


def _matlab_value(value: object) -> str:
    if value is None:
        return "NaN"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return repr(float(value))
    return "'" + str(value).replace("'", "''") + "'"


if __name__ == "__main__":
    raise SystemExit(main())
