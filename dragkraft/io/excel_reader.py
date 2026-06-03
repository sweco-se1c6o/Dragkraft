from __future__ import annotations

import math
from collections.abc import Callable
from pathlib import Path
from typing import TypeVar

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from dragkraft.domain.track import (
    CurveSegment,
    GradientSegment,
    SignalBlock,
    SpeedLimitSegment,
    Stop,
    TimingPoint,
    TrackProfile,
    TunnelSegment,
)

T = TypeVar("T")

DATA_START_ROW = 5
STRAIGHT_RADIUS_SENTINEL = 99999


class FormulaCacheError(ValueError):
    """Raised when a formula cell needed by the legacy contract has no value."""


class ExcelContractError(ValueError):
    """Raised when workbook contents do not satisfy the legacy fixed layout."""


def read_track_profile(
    workbook_path: str | Path,
    sheet_name: str,
    *,
    speed_override_kmh: float | None = None,
) -> TrackProfile:
    """Read one legacy Dragkraft worksheet without changing its Excel contract."""
    workbook_path = Path(workbook_path)
    values_workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    formulas_workbook = load_workbook(workbook_path, data_only=False, read_only=True)

    if sheet_name not in values_workbook.sheetnames:
        raise ExcelContractError(f"Workbook has no sheet named {sheet_name!r}")

    values = values_workbook[sheet_name]
    formulas = formulas_workbook[sheet_name]

    return TrackProfile(
        sheet_name=sheet_name,
        speed_limits=tuple(
            _read_rows(
                values,
                formulas,
                columns=(2, 3, 4),
                row_factory=lambda row: SpeedLimitSegment(
                    from_km=_as_float(row[0], "speed from"),
                    to_km=_as_float(row[1], "speed to"),
                    speed_kmh=(
                        float(speed_override_kmh)
                        if speed_override_kmh is not None
                        else _as_float(row[2], "speed")
                    ),
                ),
            )
        ),
        gradients=tuple(
            _read_rows(
                values,
                formulas,
                columns=(7, 8, 9),
                row_factory=lambda row: GradientSegment(
                    from_km=_as_float(row[0], "gradient from"),
                    to_km=_as_float(row[1], "gradient to"),
                    gradient_promille=_as_float(row[2], "gradient"),
                ),
            )
        ),
        tunnels=tuple(
            _read_rows(
                values,
                formulas,
                columns=(12, 13, 14),
                row_factory=lambda row: TunnelSegment(
                    from_km=_as_float(row[0], "tunnel from"),
                    to_km=_as_float(row[1], "tunnel to"),
                    factor=_as_float(row[2], "tunnel factor"),
                ),
            )
        ),
        timing_points=tuple(
            _read_rows(
                values,
                formulas,
                columns=(17, 18),
                stop_on_primary_only=True,
                row_factory=lambda row: TimingPoint(
                    position_km=_as_float(row[0], "timing point position"),
                    name=_as_text(row[1]),
                ),
            )
        ),
        stops=tuple(
            _read_rows(
                values,
                formulas,
                columns=(20, 21, 22),
                stop_on_primary_only=True,
                row_factory=lambda row: Stop(
                    position_km=_as_float(row[0], "stop position"),
                    name=_as_text(row[1]),
                    stop_time_s=_as_float(row[2], "stop time"),
                ),
            )
        ),
        curves=tuple(
            _read_rows(
                values,
                formulas,
                columns=(25, 26, 27),
                row_factory=lambda row: CurveSegment(
                    from_km=_as_float(row[0], "curve from"),
                    to_km=_as_float(row[1], "curve to"),
                    radius_m=_normalize_curve_radius(
                        _as_float(row[2], "curve radius")
                    ),
                ),
            )
        ),
        signals=tuple(
            _read_rows(
                values,
                formulas,
                columns=(30, 31, 32, 33, 34, 35),
                stop_on_primary_only=True,
                row_factory=lambda row: SignalBlock(
                    position_km=_as_float(row[0], "signal position"),
                    name=_as_text(row[1]),
                    release_speed_kmh=_as_float(row[2], "signal release speed"),
                    overlap_m=_as_float(row[3], "signal overlap"),
                    release_time_s=_as_float(row[4], "signal release time"),
                    setting_time_s=_as_float(row[5], "signal setting time"),
                ),
            )
        ),
    )


def _read_rows(
    values: Worksheet,
    formulas: Worksheet,
    *,
    columns: tuple[int, ...],
    row_factory: Callable[[tuple[object, ...]], T],
    stop_on_primary_only: bool = False,
) -> list[T]:
    records: list[T] = []
    row_number = DATA_START_ROW
    while True:
        formula_cells = [formulas.cell(row_number, column) for column in columns]
        value_cells = [values.cell(row_number, column) for column in columns]

        if _is_blank(value_cells[0].value):
            break

        for formula_cell, value_cell in zip(formula_cells, value_cells, strict=True):
            if _is_uncached_formula(formula_cell.value, value_cell.value):
                raise FormulaCacheError(
                    f"Formula cell {formula_cell.coordinate} has no cached value"
                )

        if stop_on_primary_only:
            active_cells = value_cells
        else:
            if any(_is_blank(cell.value) for cell in value_cells):
                break
            active_cells = value_cells

        records.append(row_factory(tuple(cell.value for cell in active_cells)))
        row_number += 1

    return records


def _is_uncached_formula(formula_value: object, cached_value: object) -> bool:
    return isinstance(formula_value, str) and formula_value.startswith("=") and cached_value is None


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, float) and math.isnan(value))


def _as_float(value: object, field_name: str) -> float:
    if isinstance(value, bool) or value is None:
        raise ExcelContractError(f"Expected numeric value for {field_name}")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ExcelContractError(f"Expected numeric value for {field_name}") from exc


def _as_text(value: object) -> str:
    return "" if value is None else str(value)


def _normalize_curve_radius(radius_m: float) -> float:
    if radius_m == STRAIGHT_RADIUS_SENTINEL:
        return math.inf
    return abs(radius_m)
